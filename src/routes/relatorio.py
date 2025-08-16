from flask import Blueprint, jsonify, send_file
from src.models.user import db
from src.models.produto import Produto
from src.models.contagem import Contagem
from sqlalchemy import func
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import io
import os
from datetime import datetime

relatorio_bp = Blueprint('relatorio', __name__)

@relatorio_bp.route('/relatorio/resumo', methods=['GET'])
def relatorio_resumo():
    """Gera resumo para relatório com totais por produto"""
    resumo = db.session.query(
        Produto.codigo,
        Produto.nome,
        func.sum(Contagem.quantidade).label('total_quantidade')
    ).join(Contagem).group_by(Produto.id, Produto.codigo, Produto.nome).order_by(Produto.codigo).all()
    
    resultado = []
    total_geral = 0
    
    for item in resumo:
        total_produto = item.total_quantidade or 0
        total_geral += total_produto
        resultado.append({
            'codigo': item.codigo,
            'nome': item.nome,
            'total_quantidade': total_produto
        })
    
    return jsonify({
        'produtos': resultado,
        'total_geral': total_geral
    })

@relatorio_bp.route('/relatorio/detalhado', methods=['GET'])
def relatorio_detalhado():
    """Gera relatório detalhado com todas as contagens"""
    contagens = db.session.query(
        Produto.codigo,
        Produto.nome,
        Contagem.lote,
        Contagem.validade_mes,
        Contagem.validade_ano,
        Contagem.quantidade
    ).join(Produto).order_by(Produto.codigo, Contagem.lote).all()
    
    resultado = []
    for contagem in contagens:
        resultado.append({
            'codigo': contagem.codigo,
            'nome': contagem.nome,
            'lote': contagem.lote,
            'validade_mes': contagem.validade_mes,
            'validade_ano': contagem.validade_ano,
            'quantidade': contagem.quantidade
        })
    
    return jsonify(resultado)

@relatorio_bp.route('/relatorio/pdf', methods=['GET'])
def gerar_relatorio_pdf():
    """Gera relatório em PDF"""
    try:
        # Buscar dados
        resumo_data = db.session.query(
            Produto.codigo,
            Produto.nome,
            func.sum(Contagem.quantidade).label('total_quantidade')
        ).join(Contagem).group_by(Produto.id, Produto.codigo, Produto.nome).order_by(Produto.codigo).all()
        
        detalhado_data = db.session.query(
            Produto.codigo,
            Produto.nome,
            Contagem.lote,
            Contagem.validade_mes,
            Contagem.validade_ano,
            Contagem.quantidade
        ).join(Produto).order_by(Produto.codigo, Contagem.lote).all()
        
        # Criar PDF em memória
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=18)
        
        # Estilos
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue
        )
        
        # Elementos do documento
        elements = []
        
        # Título
        title = Paragraph("Relatório de Contagem de Estoque", title_style)
        elements.append(title)
        
        # Data de geração
        data_geracao = Paragraph(f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}", styles['Normal'])
        elements.append(data_geracao)
        elements.append(Spacer(1, 20))
        
        # Resumo por produto
        elements.append(Paragraph("Resumo por Produto", heading_style))
        
        # Tabela de resumo
        resumo_table_data = [['Código', 'Produto', 'Total']]
        total_geral = 0
        
        for item in resumo_data:
            total_produto = item.total_quantidade or 0
            total_geral += total_produto
            resumo_table_data.append([
                item.codigo,
                item.nome,
                str(total_produto)
            ])
        
        # Linha de total
        resumo_table_data.append(['', 'TOTAL GERAL:', str(total_geral)])
        
        resumo_table = Table(resumo_table_data, colWidths=[1.5*inch, 3*inch, 1*inch])
        resumo_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(resumo_table)
        elements.append(Spacer(1, 30))
        
        # Detalhamento por lote
        elements.append(Paragraph("Detalhamento por Lote", heading_style))
        
        detalhado_table_data = [['Código', 'Produto', 'Lote', 'Validade', 'Quantidade']]
        
        meses = ['', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
                'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        
        for item in detalhado_data:
            validade = f"{meses[item.validade_mes]}/{item.validade_ano}"
            detalhado_table_data.append([
                item.codigo,
                item.nome[:25] + '...' if len(item.nome) > 25 else item.nome,
                item.lote,
                validade,
                str(item.quantidade)
            ])
        
        detalhado_table = Table(detalhado_table_data, colWidths=[1*inch, 2.5*inch, 1*inch, 0.8*inch, 0.7*inch])
        detalhado_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        elements.append(detalhado_table)
        
        # Construir PDF
        doc.build(elements)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'relatorio_estoque_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf',
            mimetype='application/pdf'
        )
        
    except Exception as e:
        return jsonify({'erro': f'Erro ao gerar PDF: {str(e)}'}), 500

@relatorio_bp.route('/relatorio/excel', methods=['GET'])
def gerar_relatorio_excel():
    """Gera relatório em Excel"""
    try:
        # Buscar dados
        resumo_data = db.session.query(
            Produto.codigo,
            Produto.nome,
            func.sum(Contagem.quantidade).label('total_quantidade')
        ).join(Contagem).group_by(Produto.id, Produto.codigo, Produto.nome).order_by(Produto.codigo).all()
        
        detalhado_data = db.session.query(
            Produto.codigo,
            Produto.nome,
            Contagem.lote,
            Contagem.validade_mes,
            Contagem.validade_ano,
            Contagem.quantidade
        ).join(Produto).order_by(Produto.codigo, Contagem.lote).all()
        
        # Criar workbook
        wb = Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        total_font = Font(bold=True)
        total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center')
        
        # Aba Resumo
        ws_resumo = wb.active
        ws_resumo.title = "Resumo por Produto"
        
        # Cabeçalho do resumo
        ws_resumo['A1'] = "Relatório de Contagem de Estoque - Resumo"
        ws_resumo['A1'].font = Font(bold=True, size=14)
        ws_resumo.merge_cells('A1:C1')
        
        ws_resumo['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
        ws_resumo.merge_cells('A2:C2')
        
        # Cabeçalhos da tabela
        headers_resumo = ['Código', 'Produto', 'Total']
        for col, header in enumerate(headers_resumo, 1):
            cell = ws_resumo.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Dados do resumo
        total_geral = 0
        row = 5
        for item in resumo_data:
            total_produto = item.total_quantidade or 0
            total_geral += total_produto
            
            ws_resumo.cell(row=row, column=1, value=item.codigo).border = border
            ws_resumo.cell(row=row, column=2, value=item.nome).border = border
            ws_resumo.cell(row=row, column=3, value=total_produto).border = border
            ws_resumo.cell(row=row, column=3).alignment = center_alignment
            row += 1
        
        # Total geral
        ws_resumo.cell(row=row, column=1, value="").border = border
        ws_resumo.cell(row=row, column=2, value="TOTAL GERAL:").border = border
        ws_resumo.cell(row=row, column=2).font = total_font
        ws_resumo.cell(row=row, column=2).fill = total_fill
        ws_resumo.cell(row=row, column=3, value=total_geral).border = border
        ws_resumo.cell(row=row, column=3).font = total_font
        ws_resumo.cell(row=row, column=3).fill = total_fill
        ws_resumo.cell(row=row, column=3).alignment = center_alignment
        
        # Ajustar largura das colunas
        ws_resumo.column_dimensions['A'].width = 15
        ws_resumo.column_dimensions['B'].width = 40
        ws_resumo.column_dimensions['C'].width = 12
        
        # Aba Detalhado
        ws_detalhado = wb.create_sheet("Detalhado por Lote")
        
        # Cabeçalho do detalhado
        ws_detalhado['A1'] = "Relatório de Contagem de Estoque - Detalhado"
        ws_detalhado['A1'].font = Font(bold=True, size=14)
        ws_detalhado.merge_cells('A1:E1')
        
        ws_detalhado['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
        ws_detalhado.merge_cells('A2:E2')
        
        # Cabeçalhos da tabela detalhada
        headers_detalhado = ['Código', 'Produto', 'Lote', 'Validade', 'Quantidade']
        for col, header in enumerate(headers_detalhado, 1):
            cell = ws_detalhado.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = center_alignment
        
        # Dados detalhados
        meses = ['', 'Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 
                'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
        
        row = 5
        for item in detalhado_data:
            validade = f"{meses[item.validade_mes]}/{item.validade_ano}"
            
            ws_detalhado.cell(row=row, column=1, value=item.codigo).border = border
            ws_detalhado.cell(row=row, column=2, value=item.nome).border = border
            ws_detalhado.cell(row=row, column=3, value=item.lote).border = border
            ws_detalhado.cell(row=row, column=4, value=validade).border = border
            ws_detalhado.cell(row=row, column=4).alignment = center_alignment
            ws_detalhado.cell(row=row, column=5, value=item.quantidade).border = border
            ws_detalhado.cell(row=row, column=5).alignment = center_alignment
            row += 1
        
        # Ajustar largura das colunas
        ws_detalhado.column_dimensions['A'].width = 15
        ws_detalhado.column_dimensions['B'].width = 40
        ws_detalhado.column_dimensions['C'].width = 15
        ws_detalhado.column_dimensions['D'].width = 12
        ws_detalhado.column_dimensions['E'].width = 12
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f'relatorio_estoque_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'erro': f'Erro ao gerar Excel: {str(e)}'}), 500

